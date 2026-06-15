"""Generacion de exportacion interactiva en Excel (.xlsx) de los leads."""

import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

PRIORITY_FILL = {
    "Alta": PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid"),
    "Media": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
    "Baja": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
}

WEBSITE_STATUS_FILL = {
    "sin_website": PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid"),
    "solo_red_social": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
    "website_deficiente": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
    "tiene_website": PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid"),
    "no_prospecto": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
}

COLUMNS = [
    ("Nombre", "business_name"),
    ("Nicho", "niche"),
    ("Categoria", "category"),
    ("Direccion", "address"),
    ("Ciudad", "city"),
    ("Estado", "state"),
    ("Pais", "country"),
    ("Telefono", "phone"),
    ("Email", "email"),
    ("Website", "website"),
    ("Rating", "rating"),
    ("Reviews", "reviews_count"),
    ("Maps URL", "maps_url"),
    ("Estado Negocio", "business_status"),
    ("Estado Website", "website_status"),
    ("Velocidad (ms)", "page_speed_ms"),
    ("SEO", "seo_notes"),
    ("Score", "score"),
    ("Prioridad", "priority"),
    ("Diagnostico", "diagnosis"),
    ("Estado de Contacto", "contact_status"),
    ("Fecha de busqueda", None),
]


def _search_date(lead) -> str:
    created = lead.created_at
    return created.strftime("%Y-%m-%d") if created else ""


def _write_sheet(ws, leads):
    ws.append([col_name for col_name, _ in COLUMNS])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for lead in leads:
        row = []
        for _, attr in COLUMNS:
            if attr is None:
                row.append(_search_date(lead))
            else:
                row.append(getattr(lead, attr))
        ws.append(row)

    last_col = get_column_letter(len(COLUMNS))
    last_row = len(leads) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    priority_idx = next(i for i, (name, _) in enumerate(COLUMNS, start=1) if name == "Prioridad")
    website_status_idx = next(i for i, (name, _) in enumerate(COLUMNS, start=1) if name == "Estado Website")

    for row_idx, lead in enumerate(leads, start=2):
        priority_fill = PRIORITY_FILL.get(lead.priority)
        if priority_fill:
            ws.cell(row=row_idx, column=priority_idx).fill = priority_fill

        website_fill = WEBSITE_STATUS_FILL.get(lead.website_status)
        if website_fill:
            ws.cell(row=row_idx, column=website_status_idx).fill = website_fill

    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row_idx in range(2, last_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def _write_summary(ws, leads):
    ws.append(["Nicho", "Fecha de busqueda", "Total leads", "Sin website", "Score promedio"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    groups: dict[tuple[str, str], list] = defaultdict(list)
    for lead in leads:
        groups[(lead.niche, _search_date(lead))].append(lead)

    for (niche, date), group_leads in sorted(groups.items()):
        total = len(group_leads)
        sin_website = sum(1 for l in group_leads if l.website_status == "sin_website")
        avg_score = sum(l.score for l in group_leads) / total if total else 0
        ws.append([niche, date, total, sin_website, round(avg_score, 1)])

    last_row = len(groups) + 1
    ws.auto_filter.ref = f"A1:E{last_row}"
    ws.freeze_panes = "A2"

    for col_idx in range(1, 6):
        max_len = len(str(ws.cell(row=1, column=col_idx).value))
        for row_idx in range(2, last_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def build_excel(leads) -> bytes:
    """Genera un archivo .xlsx con hojas por nicho, hoja resumen y, si aplica, hoja 'Todos'."""
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Resumen"
    _write_summary(summary_ws, leads)

    by_niche: dict[str, list] = defaultdict(list)
    for lead in leads:
        by_niche[lead.niche].append(lead)

    invalid_chars = set('[]:*?/\\')
    for niche, niche_leads in by_niche.items():
        raw_name = niche or "Sin nicho"
        sheet_name = "".join(c for c in raw_name if c not in invalid_chars)[:31] or "Sin nicho"
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, niche_leads)

    if len(by_niche) > 1:
        all_ws = wb.create_sheet(title="Todos")
        _write_sheet(all_ws, leads)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
